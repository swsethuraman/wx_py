import pandas as pd
import numpy as np
import openpyxl
from datetime import date, timedelta, datetime
import xlsxwriter
import openpyxl as op
from openpyxl import load_workbook
from openpyxl.styles import Color, PatternFill, Font, Border, Side, Alignment, Protection
from openpyxl.styles import colors
from openpyxl.cell import Cell
import os


def read_deal_tracker():
    df = pd.read_excel('/shared/wx/Pricing and Deals/Origination - Deal Tracker/Origination Deal Tracker.xlsx')
    #FILTER OUT EXCHANGE TRADED DEALS
    df = df.iloc[np.where(df['OTC/Secondary/Broker market ']=='OTC')].reset_index(drop=True)
    #FILTER OUT DEALS WITH VALID INITIATE/RESPOND VALUES
    df = df[['Date created (when JW creates Deal ID)','Initiate/Respond','Quote Provided (Y/N)','Trade Status','Regulatory Region']].iloc[np.where((df['Initiate/Respond']=='Initiate') | (df['Initiate/Respond']=='Respond'))].reset_index(drop=True)
    #FORMAT
    df = df.rename(columns={'Date created (when JW creates Deal ID)':'DATE','Initiate/Respond':'INITIATED','Quote Provided (Y/N)':'QUOTED','Trade Status':'TRADED','Regulatory Region':'REGION'})
    df['DATE'] = [datetime.date(d) for d in pd.to_datetime(df['DATE'])]
    df['MONTH'] = pd.to_datetime(df['DATE']).dt.month
    df['YEAR'] = pd.to_datetime(df['DATE']).dt.year
    df['QUARTER'] = (pd.to_datetime(df['DATE']).dt.month/4)
    df['QUARTER'] = df['QUARTER'].apply(np.ceil).astype(int)
    df['INITIATOR'] = np.where(df['INITIATED']=='Initiate','Laurion','Counterparty')
    return df

def origination_stats(df_ini,deal_count):
    deals_initiated = len(df_ini.index)
    if deal_count==0:
        initiated_per = 0
    else:
        initiated_per = 100*deals_initiated/deal_count
    deals_quoted = len(df_ini.iloc[np.where(df_ini['QUOTED']=='Quoted')].index)
    deals_traded = len(df_ini.iloc[np.where(df_ini['TRADED']=='Bound')].index)
    if deals_initiated==0:
        quoted_per = 0
        traded_per = 0
    else:
        quoted_per = 100*len(df_ini.iloc[np.where(df_ini['QUOTED']=='Quoted')].index)/deals_initiated
        traded_per = 100*len(df_ini.iloc[np.where(df_ini['TRADED']=='Bound')].index)/deals_initiated
    return deals_initiated, initiated_per, deals_quoted, quoted_per, deals_traded, traded_per

def format_stats(df):
    initiators=['Laurion','Counterparty']
    if datetime.now().month==1:
        latest_month = 12
        latest_quarter = 4
        latest_year = datetime.now().year-1
    else:
        latest_month = datetime.now().month-1
        latest_quarter = int(np.ceil((datetime.now().month-1)/4))
        latest_year = datetime.now().year
    df_final = pd.DataFrame(columns=['ORDER','TIMEFRAME','INITIATOR','DEALS INITIATED','INITIATED(%)','DEALS QUOTED','QUOTED(%)','DEALS TRADED','TRADED(%)','TOTAL QUOTES'])
    df_1m = df.iloc[np.where((df['MONTH']==latest_month)&(df['YEAR']==latest_year))].reset_index(drop=True)
    df_q = df.iloc[np.where((df['QUARTER']==latest_quarter)&(df['YEAR']==latest_year))].reset_index(drop=True)
    df_6m = df.iloc[np.where(((df['MONTH']>latest_month-6)&(df['YEAR']==latest_year))|((df['MONTH']>latest_month+6)&(df['YEAR']==latest_year-1)))].reset_index(drop=True)
    df_y = df.iloc[np.where(df['YEAR']==latest_year)].reset_index(drop=True)
    for i in initiators:
        #PAST 1 MONTH
        df_ini = df_1m.iloc[np.where(df_1m['INITIATOR']==i)].reset_index(drop=True)
        deal_count = len(df_1m.index)
        deals_initiated, initiated_per, deals_quoted, quoted_per, deals_traded, traded_per = origination_stats(df_ini,deal_count)
        total_quotes = len(df_1m.iloc[np.where(df_1m['QUOTED']=='Quoted')].index)
        new_row = {'ORDER':1,'TIMEFRAME':'Previous Month','INITIATOR':i,'DEALS INITIATED':deals_initiated,'INITIATED(%)':round(initiated_per,2),'DEALS QUOTED':deals_quoted,'QUOTED(%)':round(quoted_per,2),'DEALS TRADED':deals_traded,'TRADED(%)':round(traded_per,2),'TOTAL QUOTES':total_quotes}
        df_final = df_final.append(new_row, ignore_index=True)
        #CURRENT QUARTER
        df_ini = df_q.iloc[np.where(df_q['INITIATOR']==i)].reset_index(drop=True)
        deal_count = len(df_q.index)
        deals_initiated, initiated_per, deals_quoted, quoted_per, deals_traded, traded_per = origination_stats(df_ini,deal_count)
        total_quotes = len(df_q.iloc[np.where(df_q['QUOTED']=='Quoted')].index)
        new_row = {'ORDER':2,'TIMEFRAME':'This Quarter','INITIATOR':i,'DEALS INITIATED':deals_initiated,'INITIATED(%)':round(initiated_per,2),'DEALS QUOTED':deals_quoted,'QUOTED(%)':round(quoted_per,2),'DEALS TRADED':deals_traded,'TRADED(%)':round(traded_per,2),'TOTAL QUOTES':total_quotes}
        df_final = df_final.append(new_row, ignore_index=True)
        #PAST 6 MONTH
        df_ini = df_6m.iloc[np.where(df_6m['INITIATOR']==i)].reset_index(drop=True)
        deal_count = len(df_6m.index)
        deals_initiated, initiated_per, deals_quoted, quoted_per, deals_traded, traded_per = origination_stats(df_ini,deal_count)
        total_quotes = len(df_6m.iloc[np.where(df_6m['QUOTED']=='Quoted')].index)
        new_row = {'ORDER':3,'TIMEFRAME':'Past 6 Months','INITIATOR':i,'DEALS INITIATED':deals_initiated,'INITIATED(%)':round(initiated_per,2),'DEALS QUOTED':deals_quoted,'QUOTED(%)':round(quoted_per,2),'DEALS TRADED':deals_traded,'TRADED(%)':round(traded_per,2),'TOTAL QUOTES':total_quotes}
        df_final = df_final.append(new_row, ignore_index=True)
        #YTD
        df_ini = df_y.iloc[np.where(df_y['INITIATOR']==i)].reset_index(drop=True)
        deal_count = len(df_y.index)
        deals_initiated, initiated_per, deals_quoted, quoted_per, deals_traded, traded_per = origination_stats(df_ini,deal_count)
        total_quotes = len(df_y.iloc[np.where(df_y['QUOTED']=='Quoted')].index)
        new_row = {'ORDER':4,'TIMEFRAME':'YTD','INITIATOR':i,'DEALS INITIATED':deals_initiated,'INITIATED(%)':round(initiated_per,2),'DEALS QUOTED':deals_quoted,'QUOTED(%)':round(quoted_per,2),'DEALS TRADED':deals_traded,'TRADED(%)':round(traded_per,2),'TOTAL QUOTES':total_quotes}
        df_final = df_final.append(new_row, ignore_index=True)
    df_final = df_final.sort_values(by=['ORDER']).reset_index(drop=True)
    df_final['STYLE'] = np.where(df_final['INITIATOR']=='Laurion','INITIATE','SHOW')
    try:
        df_final['TOTAL QUOTES(%)'] = 100*df_final['DEALS QUOTED']/df_final['TOTAL QUOTES']
    except:
        df_final['TOTAL QUOTES(%)'] = 0
    df_final['TOTAL QUOTES(%)'] = df_final['TOTAL QUOTES(%)'].astype(float).round(2)
    df_final = df_final[['STYLE','TIMEFRAME','INITIATOR','DEALS INITIATED','INITIATED(%)','DEALS QUOTED','QUOTED(%)','DEALS TRADED','TRADED(%)','TOTAL QUOTES','TOTAL QUOTES(%)']]
    return df_final


def style_xlsx(current_file,region,mon,day,yr):
    #LOAD WORKSHEET
    wb = op.load_workbook(filename=current_file)
    ws = wb[region]
    #SET STATIC CELL VALUES
    ws.cell(row=2, column=2).value = 'WX ' + region + ' OTC Trading MIS'
    ws.cell(row=3, column=2).value = mon+'/'+day+'/'+yr
    ws.cell(row=4, column=4).value = 'Guardrail Definitions:'
    ws.cell(row=4, column=5).value = 'Laurion initiates 60%+ of activity'
    ws.cell(row=4, column=7).value = 'Laurion quotes/prices 40% or less of deals that counterparties show us'
    ws.cell(row=4, column=9).value = 'Laurion executes 25% or less of the trades that we quote off the back of counterparty shows'
    ws.cell(row=4, column=11).value = 'Laurion\'s quotes/prices off the back of counterparty shows account for less than 40% of total quotes'
    #CELL COLOURS
    blueFill = PatternFill(start_color='00003366',end_color='00003366',fill_type='solid')
    greenFill = PatternFill(start_color='00CCFFCC',end_color='00CCFFCC',fill_type='solid')
    pinkFill = PatternFill(start_color='00FF99CC',end_color='00FF99CC',fill_type='solid')
    orangeFill = PatternFill(start_color='00FFCC99',end_color='00FFCC99',fill_type='solid')
    purpleFill = PatternFill(start_color='00CC99FF',end_color='00CC99FF',fill_type='solid')
    yellowFill = PatternFill(start_color='00FFFF00',end_color='00FFFF00',fill_type='solid')
    rowFill = PatternFill(start_color='00CCFFFF',end_color='00CCFFFF',fill_type='solid')
    def fill_cell(cols,rows,fill):
        for r in rows:
            for c in cols:
                ws[c+r].fill = fill
    fill_cell(['E','G','I','K'],['4'],blueFill)
    fill_cell(['E','F'],['5'],greenFill)
    fill_cell(['G','H'],['5'],pinkFill)
    fill_cell(['I','J'],['5'],orangeFill)
    fill_cell(['K','L'],['5'],purpleFill)
    fill_cell(['B','C','D','E','F','G','H','I','J','K','L'],['8','9','12','13'],rowFill)
    #FONTS
    whiteFont = Font(color='00FFFFFF')
    boldFont = Font(bold=True)
    ws['E4'].font = whiteFont
    ws['G4'].font = whiteFont
    ws['I4'].font = whiteFont
    ws['K4'].font = whiteFont
    ws['B2'].font = boldFont
    ws['D4'].font = boldFont
    #WIDTHS AND HEIGHTS
    ws.column_dimensions['A'].width = 2
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 13
    ws.column_dimensions['E'].width = 17
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 16
    ws.column_dimensions['H'].width = 13
    ws.column_dimensions['I'].width = 16
    ws.column_dimensions['J'].width = 13
    ws.column_dimensions['K'].width = 16
    ws.column_dimensions['L'].width = 18
    ws.row_dimensions[4].height = 60
    #BORDERS
    thin = Side(border_style='thin', color='000000')
    thick = Side(border_style='thick', color='000000')
    def set_border(rows,cols,top,left,right,bottom):
        for r in rows:
            for c in cols:
                border = Border(top=top, left=left, right=right, bottom=bottom)
                ws[c+r].border = border
    set_border(['4'],['E','F','G','H','I','J','K','L'],thick,thick,thick,thick)
    set_border(['5'],['B','C','D','E','F','G','H','I','J','K','L'],thick,thick,thick,thick)
    set_border(['6','7','8','9','10','11','12'],['B','E','G','I','K'],thin,thick,thin,thin)
    set_border(['6','7','8','9','10','11','12'],['D','F','H','J','L'],thin,thin,thick,thin)
    set_border(['6','7','8','9','10','11','12'],['C'],thin,thin,thin,thin)
    set_border(['13'],['B','E','G','I','K'],thin,thick,thin,thick)
    set_border(['13'],['D','F','H','J','L'],thin,thin,thick,thick)
    set_border(['13'],['C'],thin,thin,thin,thick)
    #ALIGNMENT
    ws['D4'].alignment = Alignment(wrap_text=True,horizontal='center',vertical='center')
    ws['E4'].alignment = Alignment(wrap_text=True,horizontal='center',vertical='center')
    ws['G4'].alignment = Alignment(wrap_text=True,horizontal='center',vertical='center')
    ws['I4'].alignment = Alignment(wrap_text=True,horizontal='center',vertical='center')
    ws['K4'].alignment = Alignment(wrap_text=True,horizontal='center',vertical='center')
    #MERGE CELLS
    ws.merge_cells('E4:F4')
    ws.merge_cells('G4:H4')
    ws.merge_cells('I4:J4')
    ws.merge_cells('K4:L4')
    #CONDITIONAL COLOURS (THIS MIGHT NEED TO BE FIRST IF CELL COLOURS CANT BE OVERWRITTEN)
    def flag_issues(cols,rows,laurion_threshold,counterparty_threshold,checker_col='E'):
        for r in rows:
            for c in cols:
                if ws['B'+r].value=='INITIATE':
                    total = ws[checker_col+r].value+ws[checker_col+str(int(r)+1)].value
                if ws['B'+r].value=='INITIATE' and ws[c+r].value<laurion_threshold and total>0:
                    ws[c+r].fill = yellowFill
                elif ws['B'+r].value=='SHOW' and ws[c+r].value>counterparty_threshold:
                    ws[c+r].fill = yellowFill
    flag_issues(['F'],['6','7','8','9','10','11','12','13'],60,40)
    flag_issues(['H'],['6','7','8','9','10','11','12','13'],0,40)
    flag_issues(['J'],['6','7','8','9','10','11','12','13'],0,25)
    flag_issues(['L'],['6','7','8','9','10','11','12','13'],0,40)
    wb.save(filename=current_file)